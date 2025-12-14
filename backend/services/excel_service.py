"""
Excel处理服务
"""
import os
import uuid
import json
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook
import aiosqlite
from config import UPLOAD_DIR
from database import DATABASE_PATH


class ExcelService:
    """Excel文件处理服务"""
    
    @staticmethod
    def parse_excel(file_path: str) -> Dict[str, Any]:
        """
        解析Excel文件，获取所有Sheet的信息
        
        Returns:
            {
                "sheets": [
                    {
                        "name": "Sheet1",
                        "columns": ["A", "B", "C"],
                        "row_count": 100,
                        "preview": [...前5行数据...]
                    }
                ]
            }
        """
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheets_info = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 获取列名（第一行）
            columns = []
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                columns = [str(cell) if cell is not None else f"列{i+1}" 
                          for i, cell in enumerate(first_row)]
            
            # 获取行数
            row_count = sheet.max_row - 1 if sheet.max_row else 0
            
            # 获取预览数据（前5行）
            preview = []
            for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
                preview.append(list(row))
            
            sheets_info.append({
                "name": sheet_name,
                "columns": columns,
                "row_count": row_count,
                "preview": preview
            })
        
        workbook.close()
        return {"sheets": sheets_info}
    
    @staticmethod
    def read_sheet_as_dataframe(file_path: str, sheet_name: str) -> pd.DataFrame:
        """读取指定Sheet为DataFrame"""
        return pd.read_excel(file_path, sheet_name=sheet_name)
    
    @staticmethod
    def read_column(file_path: str, sheet_name: str, column_name: str) -> pd.Series:
        """读取指定列"""
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if column_name in df.columns:
            return df[column_name]
        raise ValueError(f"列 '{column_name}' 不存在于Sheet '{sheet_name}'")
    
    @staticmethod
    async def save_file_record(file_id: str, filename: str, original_name: str, 
                               file_path: str, sheets: List[Dict]) -> None:
        """保存文件记录到数据库"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            await db.execute(
                """INSERT INTO uploaded_files (id, filename, original_name, file_path, sheets)
                   VALUES (?, ?, ?, ?, ?)""",
                (file_id, filename, original_name, file_path, json.dumps(sheets, ensure_ascii=False))
            )
            await db.commit()
    
    @staticmethod
    async def get_file_record(file_id: str) -> Optional[Dict]:
        """获取文件记录"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(
                "SELECT * FROM uploaded_files WHERE id = ?", (file_id,)
            )
            row = await cursor.fetchone()
            if row:
                return dict(row)
            return None
    
    @staticmethod
    async def delete_file_record(file_id: str) -> None:
        """从数据库删除文件记录"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            await db.execute(
                "DELETE FROM uploaded_files WHERE id = ?", (file_id,)
            )
            await db.commit()
    
    @staticmethod
    async def get_all_files() -> List[Dict]:
        """获取所有上传的文件"""
        async with aiosqlite.connect(DATABASE_PATH) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(
                "SELECT * FROM uploaded_files ORDER BY created_at DESC"
            )
            rows = await cursor.fetchall()
            return [dict(row) for row in rows]
    
    @staticmethod
    def export_dataframe(df: pd.DataFrame, output_path: str) -> str:
        """导出DataFrame为Excel文件"""
        df.to_excel(output_path, index=False)
        return output_path
